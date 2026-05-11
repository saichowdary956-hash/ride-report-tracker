import argparse
import csv
import json
import os
import sqlite3
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
import re
import shutil
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


RIDE_HEADERS = [
    "Source File",
    "Full Drive ID",
    "Drive id",
    "Date",
    "RSU Startdate",
    "Driver",
    "Annotator",
    "RSU No",
    "RSU Storage %",
    "Session Starttime",
    "Session Endtime",
    "Overall Sessiontime",
    "Session Minutes",
    "Comments",
]

CATEGORY_HEADERS = [
    "Source File",
    "Full Drive ID",
    "Drive id",
    "Date",
    "RSU No",
    "Category",
    "Label",
    "Duration",
    "Minutes",
]

DAILY_HEADERS = [
    "Date",
    "Drive Count",
    "Total Session Time",
    "Total Session Minutes",
    "Avg RSU Storage %",
    "Min RSU Storage %",
    "Max RSU Storage %",
]

BURN_HEADERS = [
    "Date",
    "Drive Count",
    "Daily Session Time",
    "Daily Session Minutes",
    "Cumulative Session Time",
    "Cumulative Session Minutes",
    "Avg RSU Storage %",
    "Latest RSU Storage %",
    "Estimated Storage Burn %",
]

BY_RSU_HEADERS = [
    "RSU No",
    "Drive Count",
    "First Date",
    "Last Date",
    "Total Session Time",
    "Total Session Minutes",
    "First Storage %",
    "Latest Storage %",
    "Estimated Storage Burn %",
]

CATEGORY_TOTAL_HEADERS = [
    "Category",
    "Label",
    "Total Duration",
    "Total Minutes",
]

LOG_HEADERS = [
    "Status",
    "Source File",
    "Message",
]

DAILY_TRACKER_COLUMNS = [
    ("Date", None, None),
    ("Drive ID", None, None),
    ("Start Location", None, None),
    ("End Location", None, None),
    ("RSU Start Date", None, None),
    ("RSU No", None, None),
    ("RSU Storage %", None, None),
    ("Session Start Time", None, None),
    ("Session End Time", None, None),
    ("Overall Session Time", None, None),
    ("Sunny", "Weather", "Sunny"),
    ("Low Sun", "Weather", "Low Sun"),
    ("Cloudy", "Weather", "Cloudy"),
    ("Rain", "Weather", "Rain"),
    ("Fog", "Weather", "Fog"),
    ("Snow", "Weather", "Snow"),
    ("City (intense traffic)", "Road Type", "City"),
    ("Country", "Road Type", "Country"),
    ("Highway", "Road Type", "Highway"),
    ("Construction", "Road Type", "Construction"),
    ("Tunnel", "Road Type", "Tunnel"),
    ("Day", "Lighting", "Day"),
    ("Dawn", "Lighting", "Dawn"),
    ("Lit Night", "Lighting", "Lit Night"),
    ("Dark Night", "Lighting", "Dark Night"),
    ("Flow", "Traffic", "Flow"),
    ("Jam", "Traffic", "Jam"),
    ("3-18 mph", "Speed", "3-18mph"),
    ("19-37 mph", "Speed", "19-37mph"),
    ("38-55 mph", "Speed", "38-55mph"),
    ("56-80 mph", "Speed", "56-80mph"),
    ("81-155 mph", "Speed", "81-155mph"),
    ("Z-frame Checker", None, None),
    ("Comments", None, None),
]

BURNDOWN_ROWS = [
    ("Weather*", "Sunny", 34),
    ("Weather*", "Low Sun", 10),
    ("Weather*", "Cloudy", 28),
    ("Weather*", "Rain", 15),
    ("Weather*", "Fog", 5),
    ("Weather*", "Snow", 5),
    ("Road Type*", "City (intense traffic)", 69),
    ("Road Type*", "Country", 15),
    ("Road Type*", "Highway", 10),
    ("Road Type*", "Construction Site", 3),
    ("Road Type*", "Tunnel*", 3),
    ("Lighting", "Day", 45),
    ("Lighting", "Dawn", 10),
    ("Lighting", "Lit Night", 23),
    ("Lighting", "Dark Night", 22),
    ("Traffic", "Flow", 90),
    ("Traffic", "Jam", 10),
    ("Speed*", "5-30 km/h (3-18 mph)", 40),
    ("Speed*", "30-60 km/h (18-37 mph)", 30),
    ("Speed*", "60-90 km/h (37-55 mph)", 15),
    ("Speed*", "90-130 km/h (55-80 mph)", 8),
    ("Speed*", "130-250 km/h (80-155 mph)*", 1),
]

BURNDOWN_LOOKUP = {
    ("Weather*", "Sunny"): ("Weather", "Sunny"),
    ("Weather*", "Low Sun"): ("Weather", "Low Sun"),
    ("Weather*", "Cloudy"): ("Weather", "Cloudy"),
    ("Weather*", "Rain"): ("Weather", "Rain"),
    ("Weather*", "Fog"): ("Weather", "Fog"),
    ("Weather*", "Snow"): ("Weather", "Snow"),
    ("Road Type*", "City (intense traffic)"): ("Road Type", "City"),
    ("Road Type*", "Country"): ("Road Type", "Country"),
    ("Road Type*", "Highway"): ("Road Type", "Highway"),
    ("Road Type*", "Construction Site"): ("Road Type", "Construction"),
    ("Road Type*", "Tunnel*"): ("Road Type", "Tunnel"),
    ("Lighting", "Day"): ("Lighting", "Day"),
    ("Lighting", "Dawn"): ("Lighting", "Dawn"),
    ("Lighting", "Lit Night"): ("Lighting", "Lit Night"),
    ("Lighting", "Dark Night"): ("Lighting", "Dark Night"),
    ("Traffic", "Flow"): ("Traffic", "Flow"),
    ("Traffic", "Jam"): ("Traffic", "Jam"),
    ("Speed*", "5-30 km/h (3-18 mph)"): ("Speed", "3-18mph"),
    ("Speed*", "30-60 km/h (18-37 mph)"): ("Speed", "19-37mph"),
    ("Speed*", "60-90 km/h (37-55 mph)"): ("Speed", "38-55mph"),
    ("Speed*", "90-130 km/h (55-80 mph)"): ("Speed", "56-80mph"),
    ("Speed*", "130-250 km/h (80-155 mph)*"): ("Speed", "81-155mph"),
}

PROJECT_TARGET_HOURS = 295
DEFAULT_VEHICLE_PREFIX = "C844925"


def normalize_key(value):
    return " ".join(str(value or "").strip().split())


def parse_percent(value):
    text = str(value or "").strip().replace("%", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_date(value):
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%m/%d/%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return text


def parse_time(value):
    text = str(value or "").strip()
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            pass
    return None


def parse_duration(value):
    text = str(value or "").strip()
    if not text:
        return timedelta()
    parts = text.split(":")
    if len(parts) != 3:
        return timedelta()
    try:
        hours, minutes, seconds = [int(part) for part in parts]
    except ValueError:
        return timedelta()
    return timedelta(hours=hours, minutes=minutes, seconds=seconds)


def parse_duration_strict(value):
    text = str(value or "").strip()
    parts = text.split(":")
    if len(parts) != 3:
        raise ValueError(f"Invalid duration '{value}'")
    try:
        hours, minutes, seconds = [int(part) for part in parts]
    except ValueError as exc:
        raise ValueError(f"Invalid duration '{value}'") from exc
    return timedelta(hours=hours, minutes=minutes, seconds=seconds)


def duration_text(delta):
    total_seconds = int(round(delta.total_seconds()))
    hours, remainder = divmod(total_seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"


def duration_minutes(delta):
    return round(delta.total_seconds() / 60, 2)


def duration_seconds(value):
    return int(parse_duration(value).total_seconds())


def seconds_to_duration(seconds):
    try:
        total_seconds = int(float(seconds or 0))
    except (TypeError, ValueError):
        total_seconds = 0
    return duration_text(timedelta(seconds=total_seconds))


def display_value(value):
    if hasattr(value, "strftime"):
        return value.strftime("%m/%d/%Y")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return "" if value is None else str(value)


def context_drive_id(path):
    pattern = re.compile(r"C\d+_\d{8}_\d{6}", re.IGNORECASE)
    for part in [path.stem, *[parent.name for parent in path.parents]]:
        match = pattern.search(part)
        if match:
            return match.group(0)
    return ""


def build_full_drive_id(path, metadata, report_date):
    found = context_drive_id(path)
    if found:
        return found
    start_time = parse_time(metadata.get("Session Starttime"))
    if hasattr(report_date, "strftime") and start_time:
        return f"{DEFAULT_VEHICLE_PREFIX}_{report_date.strftime('%Y%m%d')}_{start_time.strftime('%H%M%S')}"
    return metadata.get("Drive id", "")


def safe_sheet(workbook, title, headers):
    if title in workbook.sheetnames:
        sheet = workbook[title]
    else:
        sheet = workbook.create_sheet(title)
    sheet.delete_rows(1, sheet.max_row)
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E8F0FE")
    return sheet


def style_cell(cell, fill=None, bold=False, align="center"):
    cell.font = Font(bold=bold, size=10)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    thin = Side(style="thin", color="000000")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def duration_to_excel_time(value):
    return parse_duration(value).total_seconds() / 86400


def category_index(category_rows):
    index = defaultdict(dict)
    for row in category_rows:
        key = row_key(row)
        if not key:
            continue
        index[key][(row.get("Category"), row.get("Label"))] = row.get("Duration", "00:00:00")
    return index


def category_totals(category_rows):
    totals = defaultdict(timedelta)
    daily = defaultdict(lambda: defaultdict(timedelta))
    for row in category_rows:
        lookup = (row.get("Category"), row.get("Label"))
        duration = parse_duration(row.get("Duration"))
        totals[lookup] += duration
        daily[row.get("Date")][lookup] += duration
    return totals, daily


def daily_rows_from_data(rides, category_rows):
    index = category_index(category_rows)
    rows = []
    uploaded_date = datetime.now().strftime("%m/%d/%Y")
    for ride in rides:
        key = row_key(ride)
        row = {}
        row["Source File"] = ride.get("Source File", "")
        row["Uploaded Date"] = ride.get("Uploaded Date", uploaded_date)
        for column_number, (header, category, label) in enumerate(DAILY_TRACKER_COLUMNS, start=1):
            if header == "Date":
                row[header] = display_value(ride.get("Date"))
            elif header == "Drive ID":
                row[header] = ride.get("Full Drive ID") or ride.get("Drive id") or ""
            elif header == "Start Location":
                row[header] = ride.get("Start Location", "")
            elif header == "End Location":
                row[header] = ride.get("End Location", "")
            elif header == "RSU Start Date":
                row[header] = display_value(ride.get("RSU Startdate"))
            elif header == "RSU No":
                row[header] = ride.get("RSU No") or ""
            elif header == "RSU Storage %":
                row[header] = display_value(ride.get("RSU Storage %"))
            elif header == "Session Start Time":
                row[header] = ride.get("Session Starttime") or ""
            elif header == "Session End Time":
                row[header] = ride.get("Session Endtime") or ""
            elif header == "Overall Session Time":
                row[header] = ride.get("Overall Sessiontime") or ""
            elif category:
                row[header] = index[key].get((category, label), "00:00:00")
            elif header == "Z-frame Checker":
                row[header] = ride.get("Z-frame Checker", "")
            elif header == "Comments":
                row[header] = str(ride.get("Comments") or "").strip()
        rows.append(row)
    return rows


def daily_row_id(row):
    base_id = (
        str(row.get("Drive ID") or "").strip()
        or f"{str(row.get('Source File') or '').strip()}::{str(row.get('Date') or '').strip()}::{str(row.get('Number') or '').strip()}"
        or str(datetime.now().timestamp())
    )
    return f"{row_vehicle(row)}::{base_id}"


def tracker_db_path(output_dir):
    return Path(output_dir) / "daily_tracker_backup.sqlite"


def database_url():
    value = os.environ.get("DATABASE_URL", "").strip()
    if value.upper().startswith("DATABASE_URL="):
        value = value.split("=", 1)[1].strip()
    value = value.strip("\"'")
    value = re.sub(r"\s+", "", value)
    if value.startswith("postgres://"):
        value = "postgresql://" + value[len("postgres://"):]
    return value


def using_cloud_database():
    return bool(database_url())


def allow_sqlite_fallback():
    return os.environ.get("ALLOW_SQLITE_FALLBACK", "1").strip().lower() not in {"0", "false", "no"}


DATABASE_FALLBACK_REASON = ""


def database_fallback_reason():
    return DATABASE_FALLBACK_REASON


def upsert_daily_row_sql():
    return """
        INSERT INTO daily_rows (id, position, vehicle, source_file, drive_id, data_json, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(id) DO UPDATE SET
            position = excluded.position,
            vehicle = excluded.vehicle,
            source_file = excluded.source_file,
            drive_id = excluded.drive_id,
            data_json = excluded.data_json,
            updated_at = excluded.updated_at
    """


def upsert_setting_sql():
    return """
        INSERT INTO settings (key, value, updated_at)
        VALUES (?, ?, ?)
        ON CONFLICT(key) DO UPDATE SET
            value = excluded.value,
            updated_at = excluded.updated_at
    """


class PostgresConnection:
    def __init__(self, conn):
        self.conn = conn

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, traceback):
        if exc_type:
            self.conn.rollback()
        self.conn.close()

    def execute(self, sql, params=None):
        cursor = self.conn.cursor()
        cursor.execute(sql.replace("?", "%s"), params or ())
        return cursor

    def commit(self):
        self.conn.commit()


def connect_postgres_tracker_db():
    try:
        import psycopg
        from psycopg.rows import dict_row
    except ImportError as exc:
        raise RuntimeError("DATABASE_URL is set, but psycopg is not installed. Run: pip install -r requirements.txt") from exc

    try:
        conn = psycopg.connect(database_url(), row_factory=dict_row)
    except psycopg.ProgrammingError as exc:
        raise RuntimeError(
            "DATABASE_URL is not a valid PostgreSQL connection string. In Render, set DATABASE_URL to the raw "
            "Supabase/Render Postgres URI only, with no quotes and no 'DATABASE_URL=' prefix. If using Supabase, "
            "append '?sslmode=require'."
        ) from exc
    db = PostgresConnection(conn)
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS daily_rows (
            id TEXT PRIMARY KEY,
            position INTEGER NOT NULL,
            vehicle TEXT,
            source_file TEXT,
            drive_id TEXT,
            data_json TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    db.execute("ALTER TABLE daily_rows ADD COLUMN IF NOT EXISTS vehicle TEXT")
    db.execute("ALTER TABLE daily_rows ADD COLUMN IF NOT EXISTS source_file TEXT")
    db.execute("ALTER TABLE daily_rows ADD COLUMN IF NOT EXISTS drive_id TEXT")
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS audit_log (
            id BIGSERIAL PRIMARY KEY,
            action TEXT NOT NULL,
            row_id TEXT,
            detail TEXT,
            created_at TEXT NOT NULL
        )
        """
    )
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    db.execute(
        """
        UPDATE daily_rows
        SET
            vehicle = COALESCE(NULLIF(vehicle, ''), COALESCE(NULLIF(data_json::jsonb ->> 'Vehicle', ''), 'Default')),
            source_file = COALESCE(NULLIF(source_file, ''), data_json::jsonb ->> 'Source File'),
            drive_id = COALESCE(NULLIF(drive_id, ''), data_json::jsonb ->> 'Drive ID')
        WHERE vehicle IS NULL OR source_file IS NULL OR drive_id IS NULL
        """
    )
    db.execute("CREATE INDEX IF NOT EXISTS idx_daily_rows_position ON daily_rows (position)")
    db.execute("CREATE INDEX IF NOT EXISTS idx_daily_rows_vehicle_position ON daily_rows (vehicle, position)")
    db.execute("CREATE INDEX IF NOT EXISTS idx_daily_rows_vehicle_source ON daily_rows (vehicle, source_file)")
    db.execute("CREATE INDEX IF NOT EXISTS idx_daily_rows_vehicle_drive ON daily_rows (vehicle, drive_id)")
    db.execute("CREATE INDEX IF NOT EXISTS idx_audit_log_created_at ON audit_log (created_at)")
    db.commit()
    return db


def connect_tracker_db(output_dir):
    global DATABASE_FALLBACK_REASON
    if using_cloud_database():
        try:
            return connect_postgres_tracker_db()
        except Exception as exc:
            if not allow_sqlite_fallback():
                raise
            DATABASE_FALLBACK_REASON = str(exc)
            print(f"WARNING: Could not connect to DATABASE_URL; falling back to local SQLite. Reason: {exc}")
    db_path = tracker_db_path(output_dir)
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.execute("PRAGMA temp_store=MEMORY")
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS daily_rows (
            id TEXT PRIMARY KEY,
            position INTEGER NOT NULL,
            vehicle TEXT,
            source_file TEXT,
            drive_id TEXT,
            data_json TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    existing_columns = {row["name"] for row in conn.execute("PRAGMA table_info(daily_rows)").fetchall()}
    if "vehicle" not in existing_columns:
        conn.execute("ALTER TABLE daily_rows ADD COLUMN vehicle TEXT")
    if "source_file" not in existing_columns:
        conn.execute("ALTER TABLE daily_rows ADD COLUMN source_file TEXT")
    if "drive_id" not in existing_columns:
        conn.execute("ALTER TABLE daily_rows ADD COLUMN drive_id TEXT")
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            action TEXT NOT NULL,
            row_id TEXT,
            detail TEXT,
            created_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    for item in conn.execute("SELECT id, data_json FROM daily_rows WHERE vehicle IS NULL OR source_file IS NULL OR drive_id IS NULL").fetchall():
        data = json.loads(item["data_json"])
        conn.execute(
            "UPDATE daily_rows SET vehicle = ?, source_file = ?, drive_id = ? WHERE id = ?",
            (row_vehicle(data), str(data.get("Source File") or "").strip(), str(data.get("Drive ID") or "").strip(), item["id"]),
        )
    conn.execute("CREATE INDEX IF NOT EXISTS idx_daily_rows_position ON daily_rows (position)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_daily_rows_vehicle_position ON daily_rows (vehicle, position)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_daily_rows_vehicle_source ON daily_rows (vehicle, source_file)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_daily_rows_vehicle_drive ON daily_rows (vehicle, drive_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_audit_log_created_at ON audit_log (created_at)")
    conn.commit()
    return conn


def get_setting(output_dir, key, default=""):
    with connect_tracker_db(output_dir) as conn:
        row = conn.execute("SELECT value FROM settings WHERE key = ?", (key,)).fetchone()
    return row["value"] if row else default


def set_setting(output_dir, key, value):
    now = datetime.now().isoformat(timespec="seconds")
    with connect_tracker_db(output_dir) as conn:
        conn.execute(
            upsert_setting_sql(),
            (key, str(value), now),
        )
        conn.commit()


def vehicle_list_from_settings(output_dir):
    raw_value = get_setting(output_dir, "vehicle_list", "[]")
    try:
        vehicles = json.loads(raw_value)
    except json.JSONDecodeError:
        vehicles = []
    result = []
    seen = set()
    for vehicle in vehicles:
        vehicle = str(vehicle or "").strip()
        if vehicle and vehicle not in seen:
            seen.add(vehicle)
            result.append(vehicle)
    if "Default" not in seen:
        result.insert(0, "Default")
    return result


def save_vehicle_list(output_dir, vehicles):
    result = []
    seen = set()
    for vehicle in vehicles:
        vehicle = str(vehicle or "").strip()
        if vehicle and vehicle not in seen:
            seen.add(vehicle)
            result.append(vehicle)
    if "Default" not in seen:
        result.insert(0, "Default")
    set_setting(output_dir, "vehicle_list", json.dumps(result))


def add_vehicle(output_dir, vehicle):
    vehicle = str(vehicle or "").strip()
    if not vehicle:
        return vehicle_list_from_settings(output_dir)
    vehicles = vehicle_list_from_settings(output_dir)
    if vehicle not in vehicles:
        vehicles.append(vehicle)
        save_vehicle_list(output_dir, vehicles)
    return vehicles


def remove_vehicle(output_dir, vehicle):
    vehicle = str(vehicle or "").strip()
    if not vehicle or vehicle == "Default":
        return 0
    vehicles = [item for item in vehicle_list_from_settings(output_dir) if item != vehicle]
    save_vehicle_list(output_dir, vehicles)
    rows = load_rows_from_database(output_dir)
    kept_rows = []
    deleted = 0
    for row in rows:
        row.pop("_id", None)
        if row_vehicle(row) == vehicle:
            deleted += 1
        else:
            kept_rows.append(row)
    save_rows_to_database(output_dir, kept_rows, action="remove-vehicle")
    with connect_tracker_db(output_dir) as conn:
        conn.execute("DELETE FROM settings WHERE key LIKE ?", (f"vehicle::{vehicle}::%",))
        conn.commit()
    return deleted


def database_has_rows(output_dir):
    with connect_tracker_db(output_dir) as conn:
        row = conn.execute("SELECT COUNT(*) AS count FROM daily_rows").fetchone()
    return bool(row["count"])


def row_vehicle(row):
    return str(row.get("Vehicle") or "Default").strip() or "Default"


def row_source_file(row):
    return str(row.get("Source File") or "").strip()


def row_drive_id(row):
    return str(row.get("Drive ID") or row.get("Full Drive ID") or row.get("Drive id") or "").strip()


def save_rows_to_database(output_dir, rows, action="sync", vehicle=None):
    now = datetime.now().isoformat(timespec="seconds")
    with connect_tracker_db(output_dir) as conn:
        if vehicle is None:
            conn.execute("DELETE FROM daily_rows")
        else:
            conn.execute("DELETE FROM daily_rows WHERE vehicle = ?", (vehicle,))
        for position, row in enumerate(rows, start=1):
            if vehicle is not None:
                row["Vehicle"] = vehicle
            row_id = daily_row_id(row)
            conn.execute(
                upsert_daily_row_sql(),
                (row_id, position, row_vehicle(row), row_source_file(row), row_drive_id(row), json.dumps(row, default=str), now),
            )
        conn.execute(
            "INSERT INTO audit_log (action, row_id, detail, created_at) VALUES (?, ?, ?, ?)",
            (action, None, f"{len(rows)} row(s)", now),
        )
        conn.commit()


def load_rows_from_database(output_dir, vehicle=None):
    with connect_tracker_db(output_dir) as conn:
        if vehicle is None:
            rows = conn.execute("SELECT id, data_json FROM daily_rows ORDER BY position, id").fetchall()
        else:
            rows = conn.execute("SELECT id, data_json FROM daily_rows WHERE vehicle = ? ORDER BY position, id", (vehicle,)).fetchall()
    result = []
    for row in rows:
        data = json.loads(row["data_json"])
        if vehicle is not None and row_vehicle(data) != vehicle:
            continue
        data["_id"] = row["id"]
        result.append(data)
    return result


def update_database_row(output_dir, row_id, row_data):
    now = datetime.now().isoformat(timespec="seconds")
    with connect_tracker_db(output_dir) as conn:
        current = conn.execute("SELECT position FROM daily_rows WHERE id = ?", (row_id,)).fetchone()
        position = current["position"] if current else (conn.execute("SELECT COALESCE(MAX(position), 0) + 1 AS next_pos FROM daily_rows").fetchone()["next_pos"])
        new_id = daily_row_id(row_data)
        conn.execute("DELETE FROM daily_rows WHERE id = ?", (row_id,))
        conn.execute(
            upsert_daily_row_sql(),
            (new_id, position, row_vehicle(row_data), row_source_file(row_data), row_drive_id(row_data), json.dumps(row_data, default=str), now),
        )
        conn.execute("INSERT INTO audit_log (action, row_id, detail, created_at) VALUES (?, ?, ?, ?)", ("update", new_id, "", now))
        conn.commit()


def add_database_row(output_dir, row_data):
    now = datetime.now().isoformat(timespec="seconds")
    row_id = daily_row_id(row_data)
    with connect_tracker_db(output_dir) as conn:
        position = conn.execute("SELECT COALESCE(MAX(position), 0) + 1 AS next_pos FROM daily_rows").fetchone()["next_pos"]
        conn.execute(
            upsert_daily_row_sql(),
            (row_id, position, row_vehicle(row_data), row_source_file(row_data), row_drive_id(row_data), json.dumps(row_data, default=str), now),
        )
        conn.execute("INSERT INTO audit_log (action, row_id, detail, created_at) VALUES (?, ?, ?, ?)", ("add", row_id, "", now))
        conn.commit()


def upsert_rows_to_database(output_dir, rows, action="upsert", vehicle=None):
    if not rows:
        return 0
    now = datetime.now().isoformat(timespec="seconds")
    count = 0
    with connect_tracker_db(output_dir) as conn:
        for row in rows:
            if vehicle is not None:
                row["Vehicle"] = vehicle
            row_id = daily_row_id(row)
            current = conn.execute("SELECT position FROM daily_rows WHERE id = ?", (row_id,)).fetchone()
            if current:
                position = current["position"]
            else:
                position = conn.execute("SELECT COALESCE(MAX(position), 0) + 1 AS next_pos FROM daily_rows").fetchone()["next_pos"]
            conn.execute(
                upsert_daily_row_sql(),
                (row_id, position, row_vehicle(row), row_source_file(row), row_drive_id(row), json.dumps(row, default=str), now),
            )
            count += 1
        conn.execute(
            "INSERT INTO audit_log (action, row_id, detail, created_at) VALUES (?, ?, ?, ?)",
            (action, None, f"{count} row(s)", now),
        )
        conn.commit()
    return count


def delete_database_row(output_dir, row_id):
    now = datetime.now().isoformat(timespec="seconds")
    with connect_tracker_db(output_dir) as conn:
        conn.execute("DELETE FROM daily_rows WHERE id = ?", (row_id,))
        conn.execute("INSERT INTO audit_log (action, row_id, detail, created_at) VALUES (?, ?, ?, ?)", ("delete", row_id, "", now))
        conn.commit()


def delete_database_rows_by_source_files(output_dir, source_files, vehicle=None):
    source_set = {str(source or "").strip() for source in source_files if str(source or "").strip()}
    if not source_set:
        return 0
    now = datetime.now().isoformat(timespec="seconds")
    with connect_tracker_db(output_dir) as conn:
        placeholders = ", ".join("?" for _ in source_set)
        params = list(source_set)
        if vehicle is None:
            count_row = conn.execute(f"SELECT COUNT(*) AS count FROM daily_rows WHERE source_file IN ({placeholders})", params).fetchone()
            conn.execute(f"DELETE FROM daily_rows WHERE source_file IN ({placeholders})", params)
        else:
            count_row = conn.execute(f"SELECT COUNT(*) AS count FROM daily_rows WHERE vehicle = ? AND source_file IN ({placeholders})", [vehicle] + params).fetchone()
            conn.execute(f"DELETE FROM daily_rows WHERE vehicle = ? AND source_file IN ({placeholders})", [vehicle] + params)
        deleted = count_row["count"] if count_row else 0
        conn.execute(
            "INSERT INTO audit_log (action, row_id, detail, created_at) VALUES (?, ?, ?, ?)",
            ("delete-source-files", None, f"{deleted} row(s)", now),
        )
        conn.commit()
    return deleted


def source_files_from_database(output_dir, vehicle=None):
    with connect_tracker_db(output_dir) as conn:
        if vehicle is None:
            rows = conn.execute(
                "SELECT source_file FROM daily_rows WHERE source_file IS NOT NULL AND source_file <> '' GROUP BY source_file ORDER BY MIN(position)"
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT source_file FROM daily_rows WHERE vehicle = ? AND source_file IS NOT NULL AND source_file <> '' GROUP BY source_file ORDER BY MIN(position)",
                (vehicle,),
            ).fetchall()
    return [row["source_file"] for row in rows]


def count_rows_in_database(output_dir, vehicle=None):
    with connect_tracker_db(output_dir) as conn:
        if vehicle is None:
            row = conn.execute("SELECT COUNT(*) AS count FROM daily_rows").fetchone()
        else:
            row = conn.execute("SELECT COUNT(*) AS count FROM daily_rows WHERE vehicle = ?", (vehicle,)).fetchone()
    return int(row["count"] or 0) if row else 0


def vehicles_from_database(output_dir):
    vehicles = vehicle_list_from_settings(output_dir)
    seen = set(vehicles)
    with connect_tracker_db(output_dir) as conn:
        rows = conn.execute("SELECT vehicle FROM daily_rows WHERE vehicle IS NOT NULL AND vehicle <> '' GROUP BY vehicle ORDER BY MIN(position)").fetchall()
    for row in rows:
        vehicle = row["vehicle"]
        if vehicle not in seen:
            seen.add(vehicle)
            vehicles.append(vehicle)
    save_vehicle_list(output_dir, vehicles)
    return vehicles or ["Default"]


def totals_from_daily_rows(rows):
    totals = []
    for header, category, label in DAILY_TRACKER_COLUMNS:
        if not category:
            continue
        total_seconds = sum(duration_seconds(row.get(header, "")) for row in rows)
        totals.append({"Category": category, "Field": header, "Total": seconds_to_duration(total_seconds)})
    return totals


def sync_database_from_tracker(output_dir, tracker_path, vehicle=None):
    if database_has_rows(output_dir) or not Path(tracker_path).exists():
        return
    import_tracker_workbook(output_dir, tracker_path, action="import-existing-workbook", vehicle=vehicle)


def import_tracker_workbook(output_dir, tracker_path, action="sync-excel-edits", vehicle=None):
    if not Path(tracker_path).exists():
        return []
    workbook = safe_load_workbook(tracker_path, data_only=True)
    if workbook is None:
        return []
    if "Daily Tracker" not in workbook.sheetnames:
        return []
    sheet = workbook["Daily Tracker"]
    headers = [cell.value for cell in sheet[1]]
    existing_by_id = {row.get("_id"): row for row in load_rows_from_database(output_dir, vehicle=vehicle)}
    rows = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if not any(values):
            continue
        row = {}
        for header, value in zip(headers, values):
            row[header] = display_value(value)
        existing = existing_by_id.get(daily_row_id(row))
        if existing and not row.get("Source File"):
            row["Source File"] = existing.get("Source File", "")
        if existing and not row.get("Uploaded Date"):
            row["Uploaded Date"] = existing.get("Uploaded Date", "")
        if vehicle is not None:
            row["Vehicle"] = vehicle
        rows.append(row)
    if rows:
        save_rows_to_database(output_dir, rows, action=action, vehicle=vehicle)
        try:
            build_tracker_from_daily_rows(Path(tracker_path), load_rows_from_database(output_dir, vehicle=vehicle))
        except PermissionError:
            pass
    return rows


def write_daily_tracker_sheet(workbook, rides, category_rows):
    sheet = safe_sheet(workbook, "Daily Tracker", [column[0] for column in DAILY_TRACKER_COLUMNS])
    sheet.sheet_view.showGridLines = False
    sheet.row_dimensions[1].height = 44
    index = category_index(category_rows)

    for cell in sheet[1]:
        fill = "D9EAD3"
        if cell.column in range(9, 15):
            fill = "CFE2F3"
        elif cell.column in range(15, 20):
            fill = "FCE4D6"
        elif cell.column in range(20, 24):
            fill = "E2F0D9"
        elif cell.column in range(24, 26):
            fill = "D9EAD3"
        elif cell.column in range(26, 31):
            fill = "FCE4D6"
        style_cell(cell, fill=fill, bold=True)

    for row_number, ride in enumerate(rides, start=2):
        key = row_key(ride)
        row_fill = "DDEBF7" if row_number % 2 == 0 else "E2F0D9"
        for column_number, (header, category, label) in enumerate(DAILY_TRACKER_COLUMNS, start=1):
            cell = sheet.cell(row=row_number, column=column_number)
            if header == "Date":
                cell.value = ride.get("Date")
                cell.number_format = "m/d/yyyy"
            elif header == "Drive ID":
                cell.value = ride.get("Full Drive ID") or ride.get("Drive id")
            elif header == "Start Location":
                cell.value = ride.get("Start Location", "")
            elif header == "End Location":
                cell.value = ride.get("End Location", "")
            elif header == "RSU Start Date":
                cell.value = ride.get("RSU Startdate")
                cell.number_format = "m/d/yyyy"
            elif header == "RSU No":
                cell.value = ride.get("RSU No")
            elif header == "RSU Storage %":
                cell.value = ride.get("RSU Storage %")
                cell.number_format = '0"%"'
            elif header == "Session Start Time":
                cell.value = ride.get("Session Starttime")
            elif header == "Session End Time":
                cell.value = ride.get("Session Endtime")
            elif header == "Overall Session Time":
                cell.value = ride.get("Overall Sessiontime")
            elif category:
                cell.value = duration_to_excel_time(index[key].get((category, label), "00:00:00"))
                cell.number_format = "[h]:mm:ss"
            elif header == "Z-frame Checker":
                cell.value = ride.get("Z-frame Checker", "")
            else:
                cell.value = str(ride.get("Comments") or "").strip() if header == "Comments" else ""
            if header == "Z-frame Checker":
                style_cell(cell, fill="D9EAD3")
            else:
                style_cell(cell, fill=row_fill, align="left" if header in ("Drive ID", "Start Location", "End Location", "Comments") else "center")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    checker_column = next((index for index, column in enumerate(DAILY_TRACKER_COLUMNS, start=1) if column[0] == "Z-frame Checker"), None)
    if checker_column:
        validation = DataValidation(type="list", formula1='"Passed,Failed,Repairable"', allow_blank=True)
        sheet.add_data_validation(validation)
        validation.add(f"{get_column_letter(checker_column)}2:{get_column_letter(checker_column)}1048576")
    widths = {
        1: 14,
        2: 22,
        3: 20,
        4: 20,
        5: 14,
        33: 16,
        34: 58,
    }
    for column_number in range(1, len(DAILY_TRACKER_COLUMNS) + 1):
        sheet.column_dimensions[get_column_letter(column_number)].width = widths.get(column_number, 12)


def write_daily_rows_sheet(workbook, rows):
    sheet = safe_sheet(workbook, "Daily Tracker", [column[0] for column in DAILY_TRACKER_COLUMNS])
    sheet.sheet_view.showGridLines = False
    sheet.row_dimensions[1].height = 44
    headers = [column[0] for column in DAILY_TRACKER_COLUMNS]
    for cell in sheet[1]:
        fill = "D9EAD3"
        if cell.column in range(9, 15):
            fill = "CFE2F3"
        elif cell.column in range(15, 20):
            fill = "FCE4D6"
        elif cell.column in range(20, 24):
            fill = "E2F0D9"
        elif cell.column in range(24, 26):
            fill = "D9EAD3"
        elif cell.column in range(26, 31):
            fill = "FCE4D6"
        style_cell(cell, fill=fill, bold=True)

    duration_headers = {header for header, category, _ in DAILY_TRACKER_COLUMNS if category}
    for row_number, row in enumerate(rows, start=2):
        row_fill = "DDEBF7" if row_number % 2 == 0 else "E2F0D9"
        for column_number, header in enumerate(headers, start=1):
            cell = sheet.cell(row=row_number, column=column_number)
            value = row.get(header, "")
            if header == "Date":
                parsed = parse_date(value)
                cell.value = parsed if hasattr(parsed, "strftime") else value
                cell.number_format = "m/d/yyyy"
            elif header == "RSU Start Date":
                parsed = parse_date(value)
                cell.value = parsed if hasattr(parsed, "strftime") else value
                cell.number_format = "m/d/yyyy"
            elif header == "RSU Storage %":
                cell.value = parse_percent(value)
                cell.number_format = '0"%"'
            elif header in duration_headers:
                cell.value = duration_to_excel_time(value)
                cell.number_format = "[h]:mm:ss"
            else:
                cell.value = value
            if header == "Z-frame Checker":
                checker_value = str(value or "").strip().lower()
                fill = "00FF00" if checker_value == "passed" else "FFC7CE" if checker_value == "failed" else "FFFF00" if checker_value == "repairable" else "D9EAD3"
                style_cell(cell, fill=fill)
            else:
                style_cell(cell, fill=row_fill, align="left" if header in ("Drive ID", "Start Location", "End Location", "Comments") else "center")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    checker_column = next((index for index, column in enumerate(DAILY_TRACKER_COLUMNS, start=1) if column[0] == "Z-frame Checker"), None)
    if checker_column:
        validation = DataValidation(type="list", formula1='"Passed,Failed,Repairable"', allow_blank=True)
        sheet.add_data_validation(validation)
        validation.add(f"{get_column_letter(checker_column)}2:{get_column_letter(checker_column)}1048576")
    widths = {1: 14, 2: 22, 3: 20, 4: 20, 5: 14, 33: 16, 34: 58}
    for column_number in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(column_number)].width = widths.get(column_number, 12)


def write_completed_totals_sheet(workbook, rows):
    sheet = safe_sheet(workbook, "Completed So Far", ["Category", "Field", "Total Completed"])
    for item in totals_from_daily_rows(rows):
        sheet.append([item["Category"], item["Field"], item["Total"]])
    for row in sheet.iter_rows(min_row=2):
        row[2].number_format = "[h]:mm:ss"
    for sheet_cell in sheet[1]:
        style_cell(sheet_cell, fill="B6D7A8", bold=True)
    autofit(sheet)


def build_tracker_from_daily_rows(tracker_path, rows):
    workbook = Workbook()
    workbook.remove(workbook.active)
    write_daily_rows_sheet(workbook, rows)
    write_completed_totals_sheet(workbook, rows)
    workbook.save(tracker_path)


def rebuild_tracker_from_database(output_dir, tracker_name="daily_tracker.xlsx", vehicle=None):
    output_dir = Path(output_dir)
    rows = load_rows_from_database(output_dir, vehicle=vehicle)
    tracker_path = output_dir / tracker_name
    backup_file(tracker_path)
    try:
        build_tracker_from_daily_rows(tracker_path, rows)
    except PermissionError:
        pass
    return tracker_path


def write_main_burndown_sheet(workbook, rides, category_rows):
    sheet = safe_sheet(workbook, "Main Burndown", [])
    sheet.sheet_view.showGridLines = False
    sheet["A1"] = "1) Route plans will vary based on the weather forecast for the respective day of data collection."
    sheet["A2"] = "2) There could be a gap between actual and planned hours and minutes for the respective day."
    sheet["A3"] = "3) For more details, please refer to the Daily Tracker document."
    for row in range(1, 4):
        sheet.cell(row=row, column=1).font = Font(bold=True, size=10)

    route_dates = sorted({ride.get("Date") for ride in rides if ride.get("Date")})
    start_date = route_dates[0] if route_dates else ""
    sheet["G5"] = f"Start Date: {start_date.strftime('%m/%d/%Y') if hasattr(start_date, 'strftime') else start_date}"
    sheet["G5"].font = Font(bold=True, size=12)

    header_row = 6
    subheader_row = 7
    data_start = 8
    totals, daily = category_totals(category_rows)

    sheet.merge_cells(start_row=header_row, start_column=1, end_row=header_row, end_column=1)
    sheet["A6"] = f"{PROJECT_TARGET_HOURS} Hours"
    sheet.merge_cells(start_row=header_row, start_column=2, end_row=header_row, end_column=5)
    sheet["B6"] = "Jeep Burndown"
    base_headers = ["Category", "Condition", "%", "Target Hours", "Minutes", "Completed Hours", "Remaining Hours"]
    for column_number, title in enumerate(base_headers, start=1):
        cell = sheet.cell(row=subheader_row, column=column_number, value=title)
        style_cell(cell, fill="D9EAD3" if column_number <= 2 else "C9DAF8", bold=True)
    for column_number in range(1, 8):
        style_cell(sheet.cell(row=header_row, column=column_number), fill="B6D7A8", bold=True)

    current_column = 8
    for day_index, date_value in enumerate(route_dates, start=1):
        sheet.merge_cells(start_row=header_row, start_column=current_column, end_row=header_row, end_column=current_column + 3)
        title = date_value.strftime("%m/%d/%Y") if hasattr(date_value, "strftime") else str(date_value)
        sheet.cell(row=header_row, column=current_column, value=f"{title} Day {day_index} Route {day_index}")
        for offset, title in enumerate(["Planned min", "Remaining minutes", "Actual", "Remarks"]):
            sheet.cell(row=subheader_row, column=current_column + offset, value=title)
        for column_number in range(current_column, current_column + 4):
            style_cell(sheet.cell(row=header_row, column=column_number), fill="B6D7A8", bold=True)
            style_cell(sheet.cell(row=subheader_row, column=column_number), fill="CFE2F3", bold=True)
        current_column += 4

    overall_col = current_column
    sheet.merge_cells(start_row=header_row, start_column=overall_col, end_row=header_row, end_column=overall_col + 1)
    sheet.cell(row=header_row, column=overall_col, value="Overall Week")
    sheet.cell(row=subheader_row, column=overall_col, value="Planned Hours")
    sheet.cell(row=subheader_row, column=overall_col + 1, value="Actual Hours")
    for column_number in (overall_col, overall_col + 1):
        style_cell(sheet.cell(row=header_row, column=column_number), fill="FFC000", bold=True)
        style_cell(sheet.cell(row=subheader_row, column=column_number), fill="FFC000", bold=True)

    for row_offset, (category, condition, percent) in enumerate(BURNDOWN_ROWS):
        row_number = data_start + row_offset
        lookup = BURNDOWN_LOOKUP[(category, condition)]
        target_hours = round(PROJECT_TARGET_HOURS * percent / 100, 2)
        completed_hours = round(totals[lookup].total_seconds() / 3600, 2)
        remaining_hours = round(target_hours - completed_hours, 2)
        values = [category, condition, percent / 100, target_hours, round(target_hours * 60, 0), completed_hours, remaining_hours]
        for column_number, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_number, column=column_number, value=value)
            if column_number == 3:
                cell.number_format = "0%"
            fill = "D9D9D9"
            if column_number == 4:
                fill = "F4B183"
            elif column_number == 5:
                fill = "9DC3E6"
            elif column_number == 6:
                fill = "00B0F0"
            elif column_number == 7:
                fill = "00B050" if remaining_hours >= 0 else "F4CCCC"
            style_cell(cell, fill=fill, align="left" if column_number <= 2 else "center")

        current_column = 8
        for date_value in route_dates:
            actual = daily[date_value][lookup]
            for offset, value in enumerate([0, duration_text(totals[lookup] - actual), duration_text(actual), ""]):
                cell = sheet.cell(row=row_number, column=current_column + offset, value=value)
                if offset in (1, 2):
                    cell.value = duration_to_excel_time(value)
                    cell.number_format = "[h]:mm:ss"
                style_cell(cell, fill="D9EAF7" if offset != 3 else "DDEBF7")
            current_column += 4
        sheet.cell(row=row_number, column=overall_col, value=target_hours)
        sheet.cell(row=row_number, column=overall_col + 1, value=completed_hours)
        style_cell(sheet.cell(row=row_number, column=overall_col), fill="92D050", bold=True)
        style_cell(sheet.cell(row=row_number, column=overall_col + 1), fill="92D050", bold=True)

    total_row = data_start + len(BURNDOWN_ROWS)
    sheet.cell(row=total_row, column=4, value=PROJECT_TARGET_HOURS)
    total_completed = round(sum(totals[lookup].total_seconds() for lookup in BURNDOWN_LOOKUP.values()) / 3600, 2)
    sheet.cell(row=total_row, column=6, value=total_completed)
    sheet.cell(row=total_row, column=7, value=round(PROJECT_TARGET_HOURS - total_completed, 2))
    for column_number in range(1, overall_col + 2):
        style_cell(sheet.cell(row=total_row, column=column_number), fill="FFFFFF", bold=True)

    milestone_row = total_row + 3
    milestone = [
        ("Project Milestone", ""),
        ("Target Hours", PROJECT_TARGET_HOURS),
        ("Completed hours", round(sum(parse_duration(ride.get("Overall Sessiontime")).total_seconds() for ride in rides) / 3600, 2)),
        ("Remaining hours", round(PROJECT_TARGET_HOURS - sum(parse_duration(ride.get("Overall Sessiontime")).total_seconds() for ride in rides) / 3600, 2)),
        ("", ""),
        ("Percentage of completion", ""),
        ("Percentage remaining", ""),
    ]
    completed = milestone[2][1]
    remaining = milestone[3][1]
    milestone[5] = ("Percentage of completion", round(completed / PROJECT_TARGET_HOURS * 100, 1) if PROJECT_TARGET_HOURS else 0)
    milestone[6] = ("Percentage remaining", round(remaining / PROJECT_TARGET_HOURS * 100, 1) if PROJECT_TARGET_HOURS else 0)
    for offset, (label, value) in enumerate(milestone):
        row = milestone_row + offset
        sheet.cell(row=row, column=2, value=label)
        sheet.cell(row=row, column=3, value=value)
        fill = "B6D7A8" if offset == 0 else "F4B183" if offset == 1 else "00B0F0" if offset == 2 else "00B050" if offset == 3 else "FFFFFF"
        if offset >= 5:
            fill = "00B0F0" if offset == 5 else "00B050"
        style_cell(sheet.cell(row=row, column=2), fill=fill, bold=True, align="left")
        style_cell(sheet.cell(row=row, column=3), fill="FFFFFF")

    widths = {1: 16, 2: 28, 3: 8, 4: 12, 5: 12, 6: 14, 7: 14}
    for column_number in range(1, overall_col + 2):
        sheet.column_dimensions[get_column_letter(column_number)].width = widths.get(column_number, 13)
    sheet.freeze_panes = "H8"


def autofit(sheet):
    for column_cells in sheet.columns:
        max_length = 0
        column = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            max_length = max(max_length, len(str(cell.value or "")))
        sheet.column_dimensions[column].width = min(max_length + 2, 45)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def add_burndown_chart(sheet):
    if sheet.max_row < 2:
        return

    session_chart = LineChart()
    session_chart.title = "Cumulative Session Minutes"
    session_chart.y_axis.title = "Minutes"
    session_chart.x_axis.title = "Date"
    session_data = Reference(sheet, min_col=6, min_row=1, max_row=sheet.max_row)
    dates = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
    session_chart.add_data(session_data, titles_from_data=True)
    session_chart.set_categories(dates)
    session_chart.height = 7
    session_chart.width = 14
    sheet.add_chart(session_chart, "K2")

    storage_chart = LineChart()
    storage_chart.title = "Latest RSU Storage %"
    storage_chart.y_axis.title = "Storage %"
    storage_chart.x_axis.title = "Date"
    storage_data = Reference(sheet, min_col=8, min_row=1, max_row=sheet.max_row)
    storage_chart.add_data(storage_data, titles_from_data=True)
    storage_chart.set_categories(dates)
    storage_chart.height = 7
    storage_chart.width = 14
    sheet.add_chart(storage_chart, "K18")


def add_category_chart(sheet):
    if sheet.max_row < 2:
        return

    chart = BarChart()
    chart.title = "Category Duration Minutes"
    chart.y_axis.title = "Minutes"
    chart.x_axis.title = "Category / Label"
    data = Reference(sheet, min_col=4, min_row=1, max_row=sheet.max_row)
    labels = Reference(sheet, min_col=2, min_row=2, max_row=sheet.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.height = 8
    chart.width = 16
    sheet.add_chart(chart, "F2")


def parse_report(path, allow_partial=False):
    with path.open(newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.reader(handle))

    if not rows or not rows[0] or rows[0][0].strip() != "DataLogger Report":
        raise ValueError("Not a RideReport DataLogger CSV")

    metadata = {}
    table_start = None
    for index, row in enumerate(rows):
        cells = [cell.strip() for cell in row]
        if not any(cells):
            continue
        if "Weather" in cells:
            table_start = index
            break
        if len(cells) >= 2:
            key = normalize_key(cells[0])
            value = cells[1].strip()
            if key and key != "DataLogger Report":
                metadata[key] = value

    categories = []
    if table_start is not None:
        headers = [cell.strip() for cell in rows[table_start]]
        category_columns = []
        for index in range(0, len(headers), 2):
            category = headers[index].strip() if index < len(headers) else ""
            if category:
                category_columns.append((category, index, index + 1))

        for row in rows[table_start + 1 :]:
            for category, label_index, duration_index in category_columns:
                label = row[label_index].strip() if label_index < len(row) else ""
                raw_duration = row[duration_index].strip() if duration_index < len(row) else ""
                if not label:
                    continue
                duration = parse_duration(raw_duration)
                categories.append(
                    {
                        "category": category,
                        "label": label,
                        "duration": duration,
                        "duration_text": raw_duration or "00:00:00",
                    }
                )

    required = ["Drive id", "Date", "RSU No", "Overall Sessiontime"]
    missing = [key for key in required if not str(metadata.get(key, "")).strip()]
    if missing and not allow_partial:
        raise ValueError("Missing required field(s): " + ", ".join(missing))

    report_date = parse_date(metadata.get("Date"))
    if str(metadata.get("Overall Sessiontime", "")).strip() and not allow_partial:
        session = parse_duration_strict(metadata.get("Overall Sessiontime"))
    else:
        session = parse_duration(metadata.get("Overall Sessiontime"))
    storage = parse_percent(metadata.get("RSU Storage"))
    full_drive_id = build_full_drive_id(path, metadata, report_date)

    ride = {
        "Source File": path.name,
        "Full Drive ID": full_drive_id,
        "Drive id": metadata.get("Drive id", ""),
        "Date": report_date,
        "RSU Startdate": parse_date(metadata.get("RSU Startdate")),
        "Driver": metadata.get("Driver", ""),
        "Annotator": metadata.get("Annotator", ""),
        "RSU No": metadata.get("RSU No", ""),
        "RSU Storage %": storage,
        "Session Starttime": metadata.get("Session Starttime", ""),
        "Session Endtime": metadata.get("Session Endtime", ""),
        "Overall Sessiontime": metadata.get("Overall Sessiontime", ""),
        "Session Minutes": duration_minutes(session) if metadata.get("Overall Sessiontime") else "",
        "Comments": metadata.get("Comments", ""),
    }

    return ride, categories


def discover_csvs(input_path):
    path = Path(input_path)
    if path.is_file():
        return [path]
    return sorted(path.glob("*.csv"))


def discover_ride_report_csvs(input_path):
    return [
        path for path in discover_csvs(input_path)
        if path.name.lower().startswith("ridereport_")
    ]


def quarantine_bad_workbook(path, reason):
    path = Path(path)
    if not path.exists():
        return None
    backup_dir = path.parent / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    target = backup_dir / f"{path.stem}_corrupt_{timestamp}{path.suffix}"
    try:
        shutil.move(str(path), str(target))
    except PermissionError:
        return None
    log_path = backup_dir / "corrupt_workbooks.log"
    with log_path.open("a", encoding="utf-8") as handle:
        handle.write(f"{datetime.now().isoformat(timespec='seconds')},{path.name},{target.name},{reason}\n")
    return target


def safe_load_workbook(path, **kwargs):
    try:
        return load_workbook(path, **kwargs)
    except (BadZipFile, InvalidFileException, KeyError, OSError) as exc:
        reason = str(exc)
        if "Bad CRC-32" in reason or isinstance(exc, (BadZipFile, InvalidFileException)):
            quarantine_bad_workbook(path, reason)
            return None
        raise


def load_existing_rides(tracker_path):
    if not tracker_path.exists():
        return {}
    workbook = safe_load_workbook(tracker_path)
    if workbook is None:
        return {}
    if "Rides" not in workbook.sheetnames:
        return {}
    sheet = workbook["Rides"]
    headers = [cell.value for cell in sheet[1]]
    rows = {}
    for values in sheet.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        key = row_key(row)
        if key and is_valid_ride_row(row):
            rows[key] = row
    return rows


def is_valid_ride_row(row):
    return bool(row_key(row))


def row_key(row):
    full_drive_id = str(row.get("Full Drive ID") or "").strip()
    drive_id = str(row.get("Drive id") or "").strip()
    source = str(row.get("Source File") or "").strip()
    return full_drive_id or drive_id or source


def build_tracker(tracker_path, rides, category_rows):
    workbook = Workbook()
    workbook.remove(workbook.active)

    write_daily_tracker_sheet(workbook, list(rides.values()), category_rows)

    rides_sheet = safe_sheet(workbook, "Rides", RIDE_HEADERS)
    for ride in rides.values():
        rides_sheet.append([ride.get(header, "") for header in RIDE_HEADERS])

    category_sheet = safe_sheet(workbook, "CategoryDurations", CATEGORY_HEADERS)
    for row in category_rows:
        category_sheet.append([row.get(header, "") for header in CATEGORY_HEADERS])

    for sheet in workbook.worksheets:
        autofit(sheet)
    workbook.save(tracker_path)


def backup_file(path):
    if not path.exists():
        return
    backup_dir = path.parent / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup_path = backup_dir / f"{path.stem}_backup{path.suffix}"
    try:
        shutil.copy2(path, backup_path)
    except PermissionError:
        pass


def daily_groups(rides):
    groups = defaultdict(list)
    for ride in rides:
        groups[ride.get("Date")].append(ride)
    return dict(sorted(groups.items(), key=lambda item: str(item[0])))


def sum_durations(rides):
    total = timedelta()
    for ride in rides:
        total += parse_duration(ride.get("Overall Sessiontime"))
    return total


def build_burndown(burndown_path, rides, category_rows):
    workbook = Workbook()
    workbook.remove(workbook.active)

    write_main_burndown_sheet(workbook, rides, category_rows)

    burn_sheet = safe_sheet(workbook, "BurnDown", BURN_HEADERS)
    cumulative = timedelta()
    for date_value, group in daily_groups(rides).items():
        daily_session = sum_durations(group)
        cumulative += daily_session
        storages = [float(row["RSU Storage %"]) for row in group if row.get("RSU Storage %") is not None]
        first_storage = storages[0] if storages else None
        latest_storage = storages[-1] if storages else None
        burn_sheet.append(
            [
                date_value,
                len(group),
                duration_text(daily_session),
                duration_minutes(daily_session),
                duration_text(cumulative),
                duration_minutes(cumulative),
                round(sum(storages) / len(storages), 2) if storages else "",
                latest_storage if latest_storage is not None else "",
                round(first_storage - latest_storage, 2) if first_storage is not None and latest_storage is not None else "",
            ]
        )

    rsu_sheet = safe_sheet(workbook, "ByRSU", BY_RSU_HEADERS)
    by_rsu = defaultdict(list)
    for ride in rides:
        by_rsu[str(ride.get("RSU No") or "")].append(ride)
    for rsu, group in sorted(by_rsu.items()):
        group = sorted(group, key=lambda row: str(row.get("Date") or ""))
        storages = [float(row["RSU Storage %"]) for row in group if row.get("RSU Storage %") is not None]
        first_storage = storages[0] if storages else None
        latest_storage = storages[-1] if storages else None
        total = sum_durations(group)
        rsu_sheet.append(
            [
                rsu,
                len(group),
                group[0].get("Date"),
                group[-1].get("Date"),
                duration_text(total),
                duration_minutes(total),
                first_storage if first_storage is not None else "",
                latest_storage if latest_storage is not None else "",
                round(first_storage - latest_storage, 2) if first_storage is not None and latest_storage is not None else "",
            ]
        )

    for sheet in workbook.worksheets:
        autofit(sheet)
    add_burndown_chart(workbook["BurnDown"])
    workbook.save(burndown_path)


def load_existing_categories(tracker_path):
    category_rows = []
    if not tracker_path.exists():
        return category_rows
    workbook = safe_load_workbook(tracker_path)
    if workbook is None:
        return category_rows
    if "CategoryDurations" not in workbook.sheetnames:
        return category_rows
    sheet = workbook["CategoryDurations"]
    headers = [cell.value for cell in sheet[1]]
    for values in sheet.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        if row_key(row):
            category_rows.append(row)
    return category_rows


def process_reports(input_path, output_dir, tracker_name="daily_tracker.xlsx", fresh=False, allow_partial=False, vehicle="Default"):
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    tracker_path = output_dir / tracker_name

    csv_paths = discover_csvs(input_path)
    if Path(input_path).is_dir():
        csv_paths = discover_ride_report_csvs(input_path)

    batch_rides = {}
    batch_category_rows = []
    processed = []
    skipped = []
    for csv_path in csv_paths:
        try:
            ride, categories = parse_report(csv_path, allow_partial=allow_partial)
        except Exception as exc:
            skipped.append({"file": csv_path.name, "message": str(exc)})
            continue
        key = row_key(ride)
        batch_rides[key] = ride
        for item in categories:
            batch_category_rows.append(
                {
                    "Source File": ride["Source File"],
                    "Full Drive ID": ride["Full Drive ID"],
                    "Drive id": ride["Drive id"],
                    "Date": ride["Date"],
                    "RSU No": ride["RSU No"],
                    "Category": item["category"],
                    "Label": item["label"],
                    "Duration": item["duration_text"],
                    "Minutes": duration_minutes(item["duration"]),
                }
            )
        processed.append(ride)

    new_rows = daily_rows_from_data(list(batch_rides.values()), batch_category_rows)
    for row in new_rows:
        row["Vehicle"] = vehicle
    for row in new_rows:
        row.pop("_id", None)
        row["Vehicle"] = vehicle

    if fresh:
        save_rows_to_database(output_dir, [], action="fresh-upload-clear", vehicle=vehicle)

    if not processed:
        detail = ""
        if skipped:
            detail = " Skipped: " + "; ".join(f"{item.get('file', '')}: {item.get('message', '')}" for item in skipped[:5])
        return {
            "processed": processed,
            "skipped": skipped,
            "tracker_path": tracker_path,
            "message": "No valid RideReport CSV files found." + detail,
        }

    upsert_rows_to_database(output_dir, new_rows, action="process-csv", vehicle=vehicle)
    try:
        build_tracker_from_daily_rows(tracker_path, load_rows_from_database(output_dir, vehicle=vehicle))
    except PermissionError:
        pass
    write_process_log(output_dir / "process_log.csv", processed, skipped)

    return {
        "processed": processed,
        "skipped": skipped,
        "tracker_path": tracker_path,
        "message": f"Processed {len(processed)} valid RideReport CSV file(s); skipped {len(skipped)} file(s).",
    }


def write_process_log(path, processed, skipped):
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(LOG_HEADERS)
        for ride in processed:
            writer.writerow(["Processed", ride.get("Source File", ""), f"Drive ID {ride.get('Full Drive ID') or ride.get('Drive id', '')}"])
        for item in skipped:
            writer.writerow(["Skipped", item["file"], item["message"]])


def main():
    parser = argparse.ArgumentParser(description="Convert RideReport CSV files into daily tracker and burn-down Excel files.")
    parser.add_argument("input", help="A RideReport CSV file or a folder containing RideReport CSV files.")
    parser.add_argument("--out", default="outputs", help="Folder where Excel files are written.")
    parser.add_argument("--tracker", default="daily_tracker.xlsx", help="Daily tracker workbook name.")
    parser.add_argument("--fresh", action="store_true", help="Rebuild outputs from the input instead of appending to existing rows.")
    parser.add_argument("--allow-partial", action="store_true", help="Process RideReport CSV files even when expected fields are missing.")
    args = parser.parse_args()
    result = process_reports(args.input, args.out, args.tracker, fresh=args.fresh, allow_partial=args.allow_partial)
    print(result["message"])
    for item in result["skipped"]:
        print(f"Skipped {item['file']}: {item['message']}")
    print(f"Updated {result['tracker_path'].resolve()}")


if __name__ == "__main__":
    main()
